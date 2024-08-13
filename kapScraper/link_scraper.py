import xlsxwriter
import os
from pathlib import Path
import subprocess
import tqdm
import openpyxl
import requests


def get_latest_entry_number(file) -> int:#son bildirim sayısı
	wb = openpyxl.load_workbook(file)#file take really long to open!
	ws = wb.active
	return ws.max_row


def get_son_bildirim_sayısı() -> int:
	file_path = Path(Path.cwd(), "Kap_linkleri_tüm", "Kap_linkleri.xlsx")
	if os.path.exists(file_path):
		return get_latest_entry_number(file_path)

#01.06.2024 itibariyle
son_bildirim_sayısı =  get_son_bildirim_sayısı() #excel dosyasını açıp son rakamı okumak lazım
ILK_BILDIRIM_SAYISI = 85084
row = 0
column = 0
sheet = 0



#kap.org'daki bütün bildirim linklerini bir excel dosyasına yazıdırıyor
#85084 ile başlıyor kap bildirimleri bundan önceki sayılar yok
#excel dosyasındaki her çalışma sayfası 65530 link alabiliyor, bu yüzden her 65530'uncu linkte yeni bir çalışma sayfası açılıyor


def write_links(son_bildirim_sayısı, worksheet, workbook) -> None:
	row = 0
	column = 0
	sheet_number = 0
	pbar = tqdm.tqdm(desc="Progress", total=(son_bildirim_sayısı - ILK_BILDIRIM_SAYISI))
	for i in range(ILK_BILDIRIM_SAYISI,son_bildirim_sayısı):
		url = "https://www.kap.org.tr/tr/Bildirim/"+str(i)
		worksheet.write(row,column,url)
		row += 1
		if i % 65530 == 0:
			sheet_number += 1
			worksheet = workbook.add_worksheet(str(sheet_number))
			row = 0
		pbar.update(1)



"""
def update(link_listesi, update_count):
	count = 0
	for sheet in link_listesi.worksheets():
		if sheet.max_row == 65535:
			count += 1
			break
		elif count >= len(link_listesi.worksheets()):
			sheet = link_listesi.add_worksheet(str(count))
			write_links(son_bildirim_sayısı,count,sheet,link_listesi)
		elif sheet.max_row < 65535:
			for j in range(sheet.max_row,sheet.max_row+update_count):
				url = "https://www.kap.org.tr/tr/Bildirim/"+str(j)
				sheet.write(j,0,url)
"""

def install_requirements(filename) -> None:
	with open(filename, 'r') as file:
		requirements = file.readlines()
	pbar = tqdm.tqdm(desc="Progress",total=len(requirements))
	for requirement in requirements:
		requirement = requirement.strip()
		if requirement:
			subprocess.call(['py', '-m', 'pip', 'install', "-q","-q","-q", requirement])
		pbar.update(1)


def link_check() -> None:#en son linkten başlayıp bütün linkleri kontrol etmek veya kırık olmayan ilk linkte durmak
	filepath = Path(Path.cwd(),"Kap_linkleri_tüm","Kap_linkleri.xlsx")
	excel_file = openpyxl.load_workbook(filepath)
	active_sheet = excel_file.active
	rows = list(active_sheet.iter_rows(min_row=1,max_row=active_sheet.max_row))
	rows = reversed(rows)
	row_count = 1
	for row in rows:
		status_code = requests.get(row)
		if status_code != 200:
			remove(row_count)
		row_count += 1


def remove(row) -> None:#removes the cell
	filepath = Path(Path.cwd(), "Kap_linkleri_tüm", "Kap_linkleri.xlsx")
	excel_file = openpyxl.load_workbook(filepath)
	active_sheet = excel_file.active
	for row in active_sheet:
		for cell in row:
			active_sheet.delete_rows(row[row],1)





def main(son_bildirim_sayısı) -> None:
	print("\n\t\t\t -----Link Scraper For KAP Declarations----\n")
	print("\nProgram writes the links in an excel file to be read later\n")
	print("\nDeclarations starts from 85084, which is harcoded in the Program\n")
	print("\nThe number changes for the last declaration\n")
	print("\nThe last declaration number is"+str(son_bildirim_sayısı)+"\n")
	print("\nAn update number will be added to this last number to scrape the latest links\n")
	print("\nChecking the required libraries\n")
	requirements_file = "requirements.txt"
	install_requirements(requirements_file)
	print("\nLibraries installed.\n")
	update_count = input("\nPlease enter the update count:")
	son_bildirim_sayısı += int(update_count)
	folder = Path(Path.cwd(),"Kap_linkleri_tüm")
	os.chdir(folder)
	workbook = xlsxwriter.workbook.Workbook("Kap_linkleri.xlsx")# dosyasının olup olamdığını kontrol etmek lazım openpyxl ile
	worksheet = workbook.add_worksheet(str(0))
	write_links(son_bildirim_sayısı,worksheet,workbook)
	while True:
		answer = input("\nDo you want to check the links?")
		if answer.lower() in ["yes","y"]:
			link_check()
		if answer.lower() in ["no","n"]:
			break
		else:
			print("\nInvalid input!")
	print("\nClosing the xlsx file.")
	workbook.close()


if __name__ == "__main__":
	main(son_bildirim_sayısı)


