import openpyxl
import xlsxwriter
from bs4 import BeautifulSoup
import requests as req
import time
import os.path
from pathlib import Path
import subprocess



def extract_text(text):#her bildirimde bulununan gereksiz kısımları çıkarmak için bu kısımları boş karakter ile değiştiriyoruz
    text = text.replace("KAP×KAP MobilAndroid UygulamasıIOS UygulamasıTest Chrome UygulamasıGörüntülee-Genel Kurule-ŞirketMerkezi Kayıt KuruluşuSermaye Piyasası Aracı Alım Satım BildirimiEnglishBildirim SorgularıBugün Gelen BildirimlerBeklenen BildirimlerDetaylı SorgulamaFinansal Tablo Kalem SorgulamaŞirketlerBIST ŞirketleriYatırım KuruluşlarıPortföy Yönetim ŞirketleriBağımsız Denetim KuruluşlarıDerecelendirme ŞirketleriDeğerleme ŞirketleriDiğer KAP Üyeleri ve İşlem Görmeyen ŞirketlerKAP Üyeliği Sona Eren ŞirketlerFonlarBorsa Yatırım FonlarıYatırım FonlarıEmeklilik Yatırım FonlarıOKS Emeklilik Yatırım FonlarıYabancı Yatırım FonlarıVarlık Finansman FonlarıKonut Finansman FonlarıGayrimenkul Yatırım FonlarıGirişim Sermayesi Yatırım FonlarıProje Finansman FonlarıTasfiye Olmuş FonlarKAP HakkındaMevzuat, Duyurular ve KılavuzlarSertifika İşlemleriGenel Bilgiİlgili LinklerSermaye Piyasası Aracı Alım Satım BildirimiKVKK Kapsamında Düzeltilmiş BildirimlerTüm KategorilerÖzel Durum AçıklamasıFinansal RaporDiğer BildirimlerFon Bildirimleri","")
    text = text.replace("***YazdırVeri Analiz PlatformuYatırımcı Bilgi MerkeziBorsa Günlük BülteniSPK BülteniGayrimenkul SertifikalarıBildirim SorgularıŞirketlerFonlarKAP HakkındaBugün Gelen BildirimlerBeklenen BildirimlerDetaylı SorgulamaFinansal Tablo Kalem SorgulamaBIST ŞirketleriYatırım KuruluşlarıPortföy Yönetim ŞirketleriBağımsız Denetim KuruluşlarıDerecelendirme ŞirketleriDeğerleme ŞirketleriDiğer KAP Üyeleri ve İşlem Görmeyen ŞirketlerKAP Üyeliği Sona Eren ŞirketlerBorsa Yatırım FonlarıYatırım FonlarıEmeklilik Yatırım FonlarıOKS Emeklilik Yatırım FonlarıYabancı Yatırım FonlarıVarlık Finansman FonlarıKonut Finansman FonlarıGayrimenkul Yatırım FonlarıGirişim Sermayesi Yatırım FonlarıProje Finansman FonlarıTasfiye Olmuş FonlarMevzuat, Duyurular ve KılavuzlarSertifika İşlemleriGenel Bilgiİlgili LinklerSermaye Piyasası Aracı Alım Satım BildirimiKVKK Kapsamında Düzeltilmiş Bildirimlerİlgili Bağlantılare-Genel Kurule-Yatırımcıe-Şirkete-VeriBorsa İstanbul A.Ş.VAPYukarıSite HaritasıİletişimYardımKişisel Verilerin KorunmasıTelif Hakkı ve Çekince İhbarı","")
    text = text.replace("YazdırVeri Analiz PlatformuYatırımcı Bilgi MerkeziBorsa Günlük BülteniSPK BülteniGayrimenkul SertifikalarıBildirim SorgularıŞirketlerFonlarKAP HakkındaBugün Gelen BildirimlerBeklenen BildirimlerDetaylı SorgulamaFinansal Tablo Kalem SorgulamaBIST ŞirketleriYatırım KuruluşlarıPortföy Yönetim ŞirketleriBağımsız Denetim KuruluşlarıDerecelendirme ŞirketleriDeğerleme ŞirketleriDiğer KAP Üyeleri ve İşlem Görmeyen ŞirketlerKAP Üyeliği Sona Eren ŞirketlerBorsa Yatırım FonlarıYatırım FonlarıEmeklilik Yatırım FonlarıOKS Emeklilik Yatırım FonlarıYabancı Yatırım FonlarıVarlık Finansman FonlarıKonut Finansman FonlarıGayrimenkul Yatırım FonlarıGirişim Sermayesi Yatırım FonlarıProje Finansman FonlarıTasfiye Olmuş FonlarMevzuat, Duyurular ve KılavuzlarSertifika İşlemleriGenel Bilgiİlgili LinklerSermaye Piyasası Aracı Alım Satım BildirimiKVKK Kapsamında Düzeltilmiş Bildirimlerİlgili Bağlantılare-Genel Kurule-Yatırımcıe-Şirkete-VeriBorsa İstanbul A.Ş.VAPYukarıSite HaritasıİletişimYardımKişisel Verilerin KorunmasıTelif Hakkı ve Çekince İhbarı","")
    return str(text)





def data_retrieve(workbook,text_file,file_num):#Dosyadaki linklere giderek metni çıkarıyor ve bir excel dosyasına kaydediyor
    text_file_sheet = text_file.active #Web sayfasından aldığımız metni yazacağımız dosya
    sheet = workbook.active#Linklerin bulunduğu dosya
    try:
        for i in range(1, sheet.max_row):  # Excel dosyasındaki tüm row'ları itere eden döngü
            if sheet.cell(i,2).value != "Tamamlandı." and sheet.cell(i,2).value != "Bozuk Link.": # Programın çalışma süresi uzun olduğu için her çalıştırdığımızda baştan başlamamak için metin çıkarma işleminin bittiği sıralara Tamamlandı. yazıyoruz
                url = sheet.cell(i, 1).value
                try:
                    html = req.get(url, allow_redirects=False)#Bazı linkler 302 redirect hatası veriyor
                except req.exceptions.ConnectionError as e:#bu hatayı yakalayıp linkin karşına bozuk link yazdıktan sonra devam ediyoruz
                    if str(e).__contains__("302"):
                        cell = sheet.cell(i, 2)
                        cell.value = "Bozuk Link."
                        continue
                    else:
                        pass

                soup = BeautifulSoup(html.content, "html.parser")
                cell = text_file_sheet.cell(i, 1)
                cell.value = extract_text(soup.get_text(strip="\n"))
                memcell = sheet.cell(i, 2)  # Metni kaydettiğimiz her linkin yanına Tamamlandı. yazıyoruz
                memcell.value = "Tamamlandı."

                if i % 200 == 0:  # 200 istekten sonra kap.org kısa zamanda çok fazla istek geldiği için bağlantıyı kesiyor bu yüzden 120s duruyoruz.
                    print("\n 120 saniye bekleniyor")
                    time.sleep(120)  # 120 saniye bekledikten sonra devam
                    print("\n" + str(i) + " linkten veri çekildi")
                    print("\nDevam Ediyor...\n")
     #Programı durduğumuzda veya bağlantıda hata olduğunda program o zamana kadarki ilerlemeyi kayıt edip kapanıyor
    except KeyboardInterrupt as e:
        text_file.save(str(Path(Path.cwd(),"Text_Dosyaları"))+"\\"+"Kap_metni_"+file_num+".xlsx")
        text_file.close()
        workbook.save(str(Path(Path.cwd(),"Kap Linkleri"))+"\\"+"Kap_linkleri_"+file_num+".xlsx")
        workbook.close()
        print("Program sonlandırıldı ve dosyalar kaydedildi.\n",e)
    except req.exceptions.ConnectionError as e:
        print("Bağlantı hatası\nProgram sonlanmadan dosyalar kaydedildi\n",e)
        text_file.save(str(Path(Path.cwd(), "Text_Dosyaları")) + "\\" + "Kap_metni_" + file_num + ".xlsx")
        text_file.close()
        workbook.save(str(Path(Path.cwd(), "Kap Linkleri")) + "\\" + "Kap_linkleri_" + file_num + ".xlsx")
        workbook.close()

def main():
    file_num = input("Kaçıncı dosya ile çalışmak istiyorsun?:")
    file = Path(Path.cwd(),"Text_Dosyaları","Kap_metni_"+file_num+".xlsx")
    link_file = Path(Path.cwd(),"Kap Linkleri","Kap_linkleri_"+file_num+".xlsx")
    if os.path.isfile(file):
        text_file = openpyxl.load_workbook(str(file))
    else:
        text_file = xlsxwriter.workbook.Workbook(str(file))# xlsxwriter ve openpyxl kütüphanelerinin farklı metodları var bu yüzden program ikinci kez çalıştığında hata veriyor
        text_file.close()
        text_file = openpyxl.load_workbook(str(file))# önlemek için eğer dosya yoksa xlsxwriter ile dosyayı oluşturup tekrar openpyxl ile açıyoruz
    start = time.time()
    workbook = openpyxl.load_workbook(str(link_file))
    end = time.time()
    interval = end - start
    print("\nDosya yüklendi.\nYüklenme " + str(interval) + " saniye sürdü.")
    print("\nProgramı durdurmak için Ctrl+C'ye basın.\n")
    data_retrieve(workbook, text_file, file_num)
    workbook.close()


def install_requirements(filename):
    with open(filename, 'r') as file:
        requirements = file.readlines()
    for requirement in requirements:
        requirement = requirement.strip()
        if requirement:
            subprocess.call(['py','-m','pip', 'install', requirement])


if __name__ == "__main__":
    requirements_file = "requirements.txt"
    install_requirements(requirements_file)
    print("Kütüphane yüklenmesi tamamlandı")
    main()






